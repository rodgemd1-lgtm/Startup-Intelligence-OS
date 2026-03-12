#!/usr/bin/env python3
"""Build International Payer Strategy Swim Lanes Excel workbook.

Corrected 2026-03-11: Exact SOM figures from Sanjib's sources,
exact capability names, simplified Gap x Country matrix with X marks.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

DARK_TEAL = "1B3A4B"
TEAL = "0A7B83"
ORACLE_RED = "C74634"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
GREEN_DARK = "006100"
GREEN_BG = "C6EFCE"
YELLOW_BG = "FFEB9C"
YELLOW_FONT = "9C6500"
ORANGE_BG = "FFC7CE"
ORANGE_FONT = "9C0006"
RED_BG = "FFC7CE"
RED_FONT = "9C0006"

HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill("solid", fgColor=DARK_TEAL)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
DATA_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color=DARK_TEAL)
SUBTITLE_FONT = Font(name="Arial", size=12, color=TEAL)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(cell, row_idx, is_number=False):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    cell.alignment = Alignment(vertical="center", wrap_text=True,
                               horizontal="center" if is_number else "left")


def auto_col_widths(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# ── Exact SOM figures from Sanjib's sources ─────────────────────────
SOM = {
    "UK":            42_041_160,
    "UAE":           12_039_300,
    "Saudi Arabia":  22_229_178,
    "Australia":     20_997_900,
    "New Zealand":   11_016_533,
    "Ireland":        4_938_570,
    "Sweden":         4_147_063,
    "Kuwait":         3_138_744,
    "Malaysia":       2_275_560,
    "Spain":         17_563_014,
    "Brazil":        22_752_538,
    "Austria":       15_567_552,
    "Finland":       10_376_554,
    "Germany":       51_793_245,
    "Israel":         4_626_720,
    "India":          6_837_725,
    "Colombia":       3_656_016,
    "Chile":          4_103_316,
    "Mexico":        10_483_200,
}

# ── WORKBOOK ────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1: Swim Lane Overview
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Swim Lane Overview"
ws1.sheet_properties.pageSetUpPr = None

ws1.merge_cells("A1:H1")
ws1["A1"] = "International Payer Strategy \u2014 Swim Lane Execution Plan"
ws1["A1"].font = TITLE_FONT
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("A2:H2")
ws1["A2"] = "Prepared: March 2026 | Owner: Mike Rodgers"
ws1["A2"].font = SUBTITLE_FONT
ws1["A2"].alignment = Alignment(horizontal="left", vertical="center")

# Row 4: Headers
headers = ["Swim Lane", "Week 1", "Week 2", "Week 3", "Week 4", "Week 5", "Week 6", "Week 7"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 8)

ws1.column_dimensions["A"].width = 35
for c in range(2, 9):
    ws1.column_dimensions[get_column_letter(c)].width = 18

TEAL_FILL = PatternFill("solid", fgColor=TEAL)
TEAL_LIGHT = PatternFill("solid", fgColor="B2DFDB")
RED_FILL = PatternFill("solid", fgColor=ORACLE_RED)
LANE_FONT = Font(name="Arial", bold=True, size=11, color=WHITE)
MILESTONE_FONT = Font(name="Arial", italic=True, size=9, color=ORACLE_RED)
LABEL_FONT = Font(name="Arial", bold=True, size=10, color=DARK_TEAL)

def fill_bar(ws, row, col_start, col_end, fill, label=None, font=None):
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.fill = fill
    if label:
        cell.value = label
        cell.font = font or LANE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = THIN_BORDER

# Row 5: Cross-cutting Data Scrape
ws1.cell(row=5, column=1, value="Cross-cutting: Data Scrape").font = LABEL_FONT
ws1.cell(row=5, column=1).alignment = Alignment(vertical="center")
fill_bar(ws1, 5, 2, 3, TEAL_FILL, "Primary Scrape", LANE_FONT)
fill_bar(ws1, 5, 4, 8, TEAL_LIGHT, "Ongoing Refresh", Font(name="Arial", italic=True, size=9, color=TEAL))

# Row 7: Lane 1
ws1.cell(row=7, column=1, value="Lane 1: Country Reqs & Likelihood").font = LABEL_FONT
ws1.cell(row=7, column=1).alignment = Alignment(vertical="center")
fill_bar(ws1, 7, 2, 4, TEAL_FILL, "Lane 1 Active", LANE_FONT)

# Row 8: Lane 1 milestones
ws1.cell(row=8, column=1, value="  Milestones").font = Font(name="Arial", italic=True, size=9, color="666666")
ms1 = [("Tier 1 scorecards", 2), ("Tier 2 scorecards", 3), ("Lane 1 complete", 4)]
for label, col in ms1:
    c = ws1.cell(row=8, column=col, value=label)
    c.font = MILESTONE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="FDE8E4")

# Row 10: Lane 2
ws1.cell(row=10, column=1, value="Lane 2: Product Gap Analysis").font = LABEL_FONT
ws1.cell(row=10, column=1).alignment = Alignment(vertical="center")
fill_bar(ws1, 10, 3, 5, TEAL_FILL, "Lane 2 Active", LANE_FONT)

# Row 11: Lane 2 milestones
ws1.cell(row=11, column=1, value="  Milestones").font = Font(name="Arial", italic=True, size=9, color="666666")
ms2 = [("Catherine call", 3), ("Gap matrix draft", 4), ("Build/buy/partner recs", 5)]
for label, col in ms2:
    c = ws1.cell(row=11, column=col, value=label)
    c.font = MILESTONE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="FDE8E4")

# Row 13: Lane 3
ws1.cell(row=13, column=1, value="Lane 3: Adoption & Feasibility").font = LABEL_FONT
ws1.cell(row=13, column=1).alignment = Alignment(vertical="center")
fill_bar(ws1, 13, 5, 7, TEAL_FILL, "Lane 3 Active", LANE_FONT)

# Row 14: Lane 3 milestones
ws1.cell(row=14, column=1, value="  Milestones").font = Font(name="Arial", italic=True, size=9, color="666666")
ms3 = [("Scoring framework", 5), ("Feasibility scores", 6), ("Composite ranking", 7)]
for label, col in ms3:
    c = ws1.cell(row=14, column=col, value=label)
    c.font = MILESTONE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="FDE8E4")

# Row 16: Final Deliverables
ws1.cell(row=16, column=1, value="Final Deliverables").font = LABEL_FONT
ws1.cell(row=16, column=1).alignment = Alignment(vertical="center")
fill_bar(ws1, 16, 7, 8, RED_FILL, "Final Delivery", Font(name="Arial", bold=True, size=11, color=WHITE))

# Row 17: Final milestones
ws1.cell(row=17, column=1, value="  Milestones").font = Font(name="Arial", italic=True, size=9, color="666666")
ms4 = [("Draft strategy deck", 7), ("Final deck + GTM playbook", 8)]
for label, col in ms4:
    c = ws1.cell(row=17, column=col, value=label)
    c.font = MILESTONE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor="FDE8E4")

ws1.sheet_properties.pageSetUpPr = None
ws1.page_setup.orientation = "landscape"
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 1

# ════════════════════════════════════════════════════════════════════
# SHEET 2: Lane 1 — Country Requirements
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Lane 1 \u2014 Country Requirements")

l1_headers = ["Country", "Tier", "SOM", "RFP Status", "RFP Likelihood (1-5)",
              "Key Buyer", "Budget Cycle", "Data Sovereignty Req", "Certification Req",
              "Local Hosting Req", "Priority Actions", "Owner", "Due Date"]
for i, h in enumerate(l1_headers, 1):
    ws2.cell(row=1, column=i, value=h)
style_header_row(ws2, 1, len(l1_headers))

countries = [
    ["UK", 1, SOM["UK"], "Active (FDP)", 5, "NHS England/NHSE Digital", "Apr fiscal year", "Yes (UK GDPR)", "NHS DTAC", "Required", "Map FDP reqs to AIDP", "Mike+Bharat", "Week 2"],
    ["UAE", 1, SOM["UAE"], "Anticipated", 3, "MoH & Prevention/Dubai Health", "Calendar year", "Yes (PDPL)", "TBD", "Preferred", "UAE procurement portal scan", "Mike", "Week 2"],
    ["Saudi Arabia", 1, SOM["Saudi Arabia"], "Anticipated", 3, "Saudi Health Council/NPHIES", "Hijri calendar", "Yes (PDPL)", "CITC cert", "Required", "Map to Vision 2030", "Mike", "Week 2"],
    ["Australia", 1, SOM["Australia"], "Medium", 3, "ADHA", "Jul fiscal year", "Yes (Privacy Act)", "TBD", "Preferred", "Interop assessment", "Mike", "Week 3"],
    ["New Zealand", 1, SOM["New Zealand"], "Active", 5, "Health NZ (Te Whatu Ora)", "Jul fiscal year", "Yes (Privacy Act)", "TBD", "Cloud-first", "Respond to active RFP", "Bharat", "Week 1"],
    ["Ireland", 2, SOM["Ireland"], "Anticipated", 2, "HSE", "Calendar year", "Yes (GDPR)", "TBD", "EU hosting", "Post-UK follow-on", "Mike", "Week 3"],
    ["Sweden", 2, SOM["Sweden"], "Low", 1, "SKR (Regions)", "Calendar year", "Yes (GDPR)", "TBD", "EU hosting", "Regional approach needed", "Mike", "Week 3"],
    ["Kuwait", 2, SOM["Kuwait"], "Low", 2, "Ministry of Health", "Calendar year", "Minimal", "TBD", "GCC hosting", "GCC alignment", "Mike", "Week 3"],
    ["Malaysia", 2, SOM["Malaysia"], "Low", 1, "MoH Malaysia", "Calendar year", "Yes (PDPA)", "TBD", "APAC hosting", "Digital health assessment", "Mike", "Week 3"],
    ["Spain", 2, SOM["Spain"], "Low", 1, "Comunidades Autonomas", "Calendar year", "Yes (GDPR)", "TBD", "EU hosting", "Regional procurement mapping", "Mike", "Week 3"],
    ["Brazil", 2, SOM["Brazil"], "Low", 2, "SUS", "Calendar year", "Yes (LGPD)", "ANVISA", "Required", "Market entry complexity assessment", "Mike", "Week 3"],
    ["Austria", 2, SOM["Austria"], "Low", 1, "ELGA", "Calendar year", "Yes (GDPR)", "TBD", "EU hosting", "EHDS alignment", "Mike", "Week 3"],
    ["Finland", 2, SOM["Finland"], "Low", 1, "Kanta services", "Calendar year", "Yes (GDPR)", "TBD", "EU hosting", "Kanta integration assessment", "Mike", "Week 3"],
    ["Germany", 3, SOM["Germany"], "Low", 1, "gematik/KVs", "Calendar year", "Yes (GDPR+BDSG)", "gematik cert", "Required", "Bavaria HDI OPPT $3M", "Mike", "Week 4"],
    ["Israel", 3, SOM["Israel"], "Low", 1, "MoH", "Calendar year", "Yes (PPL)", "TBD", "Preferred", "4 HMOs dominate market", "Mike", "Week 4"],
    ["India", 3, SOM["India"], "Low", 1, "NHA (Ayushman Bharat)", "Apr fiscal year", "Yes (DPDP)", "TBD", "Required", "Scale vs per-capita analysis", "Mike", "Week 4"],
    ["Colombia", 3, SOM["Colombia"], "Low", 1, "ADRES/MoH", "Calendar year", "Yes (Law 1581)", "TBD", "Preferred", "Limited digital health infra", "Mike", "Week 4"],
    ["Chile", 3, SOM["Chile"], "Low", 1, "FONASA", "Calendar year", "Yes (Law 19628)", "TBD", "Preferred", "Small market proof-point only", "Mike", "Week 4"],
    ["Mexico", 3, SOM["Mexico"], "Low", 1, "IMSS", "Calendar year", "Yes (LFPDPPP)", "TBD", "Required", "Budget volatility risk", "Mike", "Week 4"],
]

for r_idx, row_data in enumerate(countries, 2):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        is_num = c_idx in (2, 3, 5)
        style_data_cell(cell, r_idx, is_number=is_num)
        if c_idx == 3:
            cell.number_format = '$#,##0'

# Totals row
total_row = len(countries) + 2
ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=total_row, column=1).border = THIN_BORDER
ws2.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
ws2.cell(row=total_row, column=3).font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=total_row, column=3).number_format = '$#,##0'
ws2.cell(row=total_row, column=3).border = THIN_BORDER
ws2.cell(row=total_row, column=3).fill = PatternFill("solid", fgColor="D5E8D4")

# Conditional formatting for RFP Likelihood (column E)
ws2.conditional_formatting.add(f"E2:E{total_row-1}",
    CellIsRule(operator="equal", formula=["5"], fill=PatternFill("solid", fgColor="006100"),
              font=Font(color=WHITE)))
ws2.conditional_formatting.add(f"E2:E{total_row-1}",
    CellIsRule(operator="equal", formula=["4"], fill=PatternFill("solid", fgColor="00B050"),
              font=Font(color=WHITE)))
ws2.conditional_formatting.add(f"E2:E{total_row-1}",
    CellIsRule(operator="equal", formula=["3"], fill=PatternFill("solid", fgColor=YELLOW_BG),
              font=Font(color=YELLOW_FONT)))
ws2.conditional_formatting.add(f"E2:E{total_row-1}",
    CellIsRule(operator="equal", formula=["2"], fill=PatternFill("solid", fgColor="FFC000"),
              font=Font(color="7F6000")))
ws2.conditional_formatting.add(f"E2:E{total_row-1}",
    CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor=RED_BG),
              font=Font(color=RED_FONT)))

ws2.freeze_panes = "A2"
auto_col_widths(ws2, min_width=12)
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["F"].width = 30
ws2.column_dimensions["K"].width = 35
ws2.page_setup.orientation = "landscape"
ws2.page_setup.fitToWidth = 1

# ════════════════════════════════════════════════════════════════════
# SHEET 3: Lane 2 — Product Gap Analysis
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Lane 2 \u2014 Product Gap Analysis")

# Table A: Gap Overview — exact capability names from Sanjib
gap_headers_a = ["Gap", "Revenue at Risk", "Build Complexity", "Recommendation", "Est. LOE (months)", "Priority"]
for i, h in enumerate(gap_headers_a, 1):
    ws3.cell(row=1, column=i, value=h)
style_header_row(ws3, 1, len(gap_headers_a))

# Exact names: "Predictive Population Health Planning", "Budget Optimisation" (British),
# "Policy & Reform Impact Assurance", "Fraud, Waste and Abuse Detection"
gaps = [
    ["Predictive Population Health Planning", "$80M+", "HIGH", "PARTNER", "12+", 1],
    ["Demand & Capacity Forecasting", "$60M+", "MEDIUM", "BUILD", "6-9", 2],
    ["Outcomes Benchmarking", "$50M+", "MEDIUM", "BUILD", "6-9", 3],
    ["Budget Optimisation", "$70M+", "HIGH", "PARTNER", "12+", 4],
    ["Policy & Reform Impact Assurance", "$40M+", "HIGH", "DEFER", "18+", 6],
    ["Fraud, Waste and Abuse Detection", "$30M+", "MEDIUM", "BUILD", "6-9", 5],
]
for r_idx, row_data in enumerate(gaps, 2):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        style_data_cell(cell, r_idx, is_number=(c_idx == 6))

# Table B: Gap x Country Matrix — simplified with X marks
# Short gap labels for column headers
GAP_SHORT = [
    "Pred Pop Health",
    "Demand & Cap",
    "Outcomes Bench",
    "Budget Optim",
    "Policy & Reform",
    "FWA Detection",
]

matrix_start = 10
ws3.cell(row=matrix_start - 1, column=1, value="Gap \u00d7 Country Matrix").font = Font(name="Arial", bold=True, size=13, color=DARK_TEAL)

matrix_headers = ["Country", "Tier", "SOM"] + GAP_SHORT
for i, h in enumerate(matrix_headers, 1):
    ws3.cell(row=matrix_start, column=i, value=h)
style_header_row(ws3, matrix_start, len(matrix_headers))

# Gap assignments per the user's spec:
# UK: All 6         UAE: 1,2,3,4       KSA: 1,2,3,4
# Australia: 1,2,3  NZ: 1,2,3          Ireland: 1,2,3,4
# Tier 2/3 others: 1,2 minimum (plus extras where relevant)
#
# Columns: Pred Pop Health(1), Demand & Cap(2), Outcomes Bench(3),
#           Budget Optim(4), Policy & Reform(5), FWA Detection(6)

def gaps_row(g1="", g2="", g3="", g4="", g5="", g6=""):
    return [g1, g2, g3, g4, g5, g6]

X = "X"
matrix_data = [
    ["UK",           1, SOM["UK"],           X, X, X, X, X, X],
    ["UAE",          1, SOM["UAE"],           X, X, X, X, "", ""],
    ["Saudi Arabia", 1, SOM["Saudi Arabia"],  X, X, X, X, "", ""],
    ["Australia",    1, SOM["Australia"],     X, X, X, "", "", ""],
    ["New Zealand",  1, SOM["New Zealand"],   X, X, X, "", "", ""],
    ["Ireland",      2, SOM["Ireland"],       X, X, X, X, "", ""],
    ["Sweden",       2, SOM["Sweden"],        X, X, "", "", "", ""],
    ["Kuwait",       2, SOM["Kuwait"],        X, X, "", "", "", ""],
    ["Malaysia",     2, SOM["Malaysia"],      X, X, "", "", "", ""],
    ["Spain",        2, SOM["Spain"],         X, X, "", "", "", ""],
    ["Brazil",       2, SOM["Brazil"],        X, X, "", "", "", ""],
    ["Austria",      2, SOM["Austria"],       X, X, "", "", "", ""],
    ["Finland",      2, SOM["Finland"],       X, X, "", "", "", ""],
    ["Germany",      3, SOM["Germany"],       X, X, "", "", "", ""],
    ["Israel",       3, SOM["Israel"],        X, X, "", "", "", ""],
    ["India",        3, SOM["India"],         X, X, "", "", "", ""],
    ["Colombia",     3, SOM["Colombia"],      X, X, "", "", "", ""],
    ["Chile",        3, SOM["Chile"],         X, X, "", "", "", ""],
    ["Mexico",       3, SOM["Mexico"],        X, X, "", "", "", ""],
]

X_FILL = PatternFill("solid", fgColor="C6EFCE")
X_FONT = Font(name="Arial", size=10, bold=True, color="006100")

for r_idx, row_data in enumerate(matrix_data, matrix_start + 1):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        is_num = c_idx in (2, 3)
        style_data_cell(cell, r_idx, is_number=is_num)
        if c_idx == 3:
            cell.number_format = '$#,##0'
        if val == "X":
            cell.fill = X_FILL
            cell.font = X_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws3.freeze_panes = "A2"
auto_col_widths(ws3, min_width=14)
ws3.column_dimensions["A"].width = 32
ws3.page_setup.orientation = "landscape"
ws3.page_setup.fitToWidth = 1

# ════════════════════════════════════════════════════════════════════
# SHEET 4: Lane 3 — Feasibility Scoring
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Lane 3 \u2014 Feasibility Scoring")

l3_headers = ["Country", "Tier", "SOM", "Regulatory (25%)", "Technical (20%)",
              "Buyer Maturity (20%)", "Partner Ecosystem (15%)", "Oracle Presence (10%)",
              "Political Stability (10%)", "Weighted Score", "Rank", "Composite Score"]
for i, h in enumerate(l3_headers, 1):
    ws4.cell(row=1, column=i, value=h)
style_header_row(ws4, 1, len(l3_headers))

feas_data = [
    ["UK",           1, SOM["UK"],           5, 5, 5, 5, 4, 4],
    ["UAE",          1, SOM["UAE"],           4, 4, 4, 3, 4, 3],
    ["Saudi Arabia", 1, SOM["Saudi Arabia"],  3, 4, 4, 3, 4, 3],
    ["Australia",    1, SOM["Australia"],     4, 5, 4, 4, 3, 5],
    ["New Zealand",  1, SOM["New Zealand"],   4, 4, 4, 3, 3, 5],
    ["Ireland",      2, SOM["Ireland"],       4, 4, 3, 3, 3, 4],
    ["Sweden",       2, SOM["Sweden"],        5, 5, 3, 2, 2, 5],
    ["Kuwait",       2, SOM["Kuwait"],        3, 3, 3, 2, 3, 3],
    ["Malaysia",     2, SOM["Malaysia"],      3, 3, 3, 2, 2, 3],
    ["Spain",        2, SOM["Spain"],         4, 4, 3, 3, 3, 4],
    ["Brazil",       2, SOM["Brazil"],        2, 3, 3, 3, 4, 2],
    ["Austria",      2, SOM["Austria"],       5, 4, 3, 3, 2, 5],
    ["Finland",      2, SOM["Finland"],       5, 5, 3, 2, 2, 5],
    ["Germany",      3, SOM["Germany"],       5, 5, 3, 4, 3, 4],
    ["Israel",       3, SOM["Israel"],        4, 4, 3, 3, 2, 3],
    ["India",        3, SOM["India"],         2, 2, 2, 3, 3, 2],
    ["Colombia",     3, SOM["Colombia"],      2, 2, 2, 2, 2, 3],
    ["Chile",        3, SOM["Chile"],         3, 3, 2, 2, 2, 4],
    ["Mexico",       3, SOM["Mexico"],        2, 3, 2, 3, 3, 2],
]

last_row = len(feas_data) + 1  # data starts row 2

for r_idx, row_data in enumerate(feas_data, 2):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r_idx, column=c_idx, value=val)
        is_num = c_idx >= 2
        style_data_cell(cell, r_idx, is_number=is_num)
        if c_idx == 3:
            cell.number_format = '$#,##0'
        elif c_idx in (4, 5, 6, 7, 8, 9):
            cell.number_format = '0'

    r = r_idx
    # J: Weighted Score = D*0.25 + E*0.20 + F*0.20 + G*0.15 + H*0.10 + I*0.10
    ws4.cell(row=r, column=10, value=f"=D{r}*0.25+E{r}*0.2+F{r}*0.2+G{r}*0.15+H{r}*0.1+I{r}*0.1")
    ws4.cell(row=r, column=10).number_format = '0.0'
    style_data_cell(ws4.cell(row=r, column=10), r, is_number=True)

    # K: Rank
    ws4.cell(row=r, column=11, value=f"=RANK(J{r},J$2:J${last_row})")
    ws4.cell(row=r, column=11).number_format = '0'
    style_data_cell(ws4.cell(row=r, column=11), r, is_number=True)

    # L: Composite Score = SOM * Weighted Score
    ws4.cell(row=r, column=12, value=f"=C{r}*J{r}")
    ws4.cell(row=r, column=12).number_format = '$#,##0'
    style_data_cell(ws4.cell(row=r, column=12), r, is_number=True)

# Conditional formatting for Weighted Score (col J)
ws4.conditional_formatting.add(f"J2:J{last_row}",
    CellIsRule(operator="greaterThanOrEqual", formula=["4"],
              fill=PatternFill("solid", fgColor="C6EFCE"),
              font=Font(color="006100")))
ws4.conditional_formatting.add(f"J2:J{last_row}",
    CellIsRule(operator="between", formula=["3", "3.9"],
              fill=PatternFill("solid", fgColor=YELLOW_BG),
              font=Font(color=YELLOW_FONT)))
ws4.conditional_formatting.add(f"J2:J{last_row}",
    CellIsRule(operator="lessThan", formula=["3"],
              fill=PatternFill("solid", fgColor=RED_BG),
              font=Font(color=RED_FONT)))

ws4.freeze_panes = "A2"
auto_col_widths(ws4, min_width=14)
ws4.column_dimensions["A"].width = 16
ws4.column_dimensions["C"].width = 18
ws4.page_setup.orientation = "landscape"
ws4.page_setup.fitToWidth = 1

# ── SAVE ────────────────────────────────────────────────────────────
output_path = "/Users/mikerodgers/Desktop/International_Payer_Strategy_Package/Swim_Lanes_Execution.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
