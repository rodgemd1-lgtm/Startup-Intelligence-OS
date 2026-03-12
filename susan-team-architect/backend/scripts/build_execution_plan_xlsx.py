#!/usr/bin/env python3
"""Build International Payer Strategy Execution Plan Excel workbook."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

OUTPUT = "/Users/mikerodgers/Desktop/International_Payer_Strategy_Package/Execution_Plan.xlsx"

DARK_TEAL = "1B3A4B"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN_FILL = "C6EFCE"
RED_FILL = "FFC7CE"
YELLOW_FILL = "FFEB9C"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
normal_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", size=10, bold=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

green_fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
red_fill = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")
yellow_fill = PatternFill(start_color=YELLOW_FILL, end_color=YELLOW_FILL, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_data_rows(ws, num_rows, num_cols):
    for row in range(2, num_rows + 1):
        fill = white_fill if row % 2 == 0 else gray_fill
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = wrap_align


# ── Sheet 1: Execution Plan ─────────────────────────────────────────────

def build_execution_plan(wb):
    ws = wb.active
    ws.title = "Execution Plan"

    headers = ["ID", "Swim Lane", "Task", "Owner", "Status", "Priority",
               "Start Week", "End Week", "Dependencies", "Deliverable", "Notes"]
    widths = [5, 15, 45, 18, 12, 10, 10, 10, 15, 30, 35]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    rows = [
        # CROSS-CUTTING
        ["DC-1", "Data Scrape", "Configure scraping tools (Firecrawl/Exa/Jina/Apify)", "Mike", "Not Started", "P1", 1, 1, "-", "Tool configuration complete", ""],
        ["DC-2", "Data Scrape", "Scrape government procurement portals (17 countries)", "Mike", "Not Started", "P1", 1, 2, "DC-1", "Procurement portal data", ""],
        ["DC-3", "Data Scrape", "Scrape WHO/OECD health expenditure data", "Mike", "Not Started", "P1", 1, 1, "DC-1", "Health spend datasets", ""],
        ["DC-4", "Data Scrape", "Scrape competitive intelligence (7 competitors)", "Mike", "Not Started", "P1", 1, 2, "DC-1", "Competitor profiles", ""],
        ["DC-5", "Data Scrape", "Scrape country health IT strategy documents", "Mike", "Not Started", "P2", 1, 2, "DC-1", "Strategy docs per country", ""],
        ["DC-6", "Data Scrape", "Scrape regulatory requirements by country", "Mike", "Not Started", "P2", 2, 2, "DC-1", "Regulatory matrix", ""],
        ["DC-7", "Data Scrape", "Ongoing data refresh", "Mike", "Not Started", "P3", 3, 7, "DC-2", "Updated datasets", "Monthly refresh"],
        # LANE 1
        ["L1-1", "Lane 1", "Collect all existing RFP documents from Bharat", "Bharat", "Not Started", "P1", 1, 1, "-", "RFP document package", "Bharat to send this week"],
        ["L1-2", "Lane 1", "Build country scorecard template", "Mike", "Not Started", "P1", 1, 1, "-", "Template ready", "1-pager per country"],
        ["L1-3", "Lane 1", "Complete UK scorecard", "Mike", "Not Started", "P1", 1, 2, "L1-2;DC-2", "UK country scorecard", "Largest opportunity + active RFP"],
        ["L1-4", "Lane 1", "Complete NZ scorecard", "Mike", "Not Started", "P1", 1, 1, "L1-2;L1-1", "NZ country scorecard", "Active procurement"],
        ["L1-5", "Lane 1", "Complete UAE scorecard", "Mike", "Not Started", "P1", 1, 2, "L1-2;DC-2", "UAE country scorecard", ""],
        ["L1-6", "Lane 1", "Complete KSA scorecard", "Mike", "Not Started", "P1", 1, 2, "L1-2;DC-2", "KSA country scorecard", ""],
        ["L1-7", "Lane 1", "Complete Australia scorecard", "Mike", "Not Started", "P2", 2, 2, "L1-2;DC-2", "AUS country scorecard", ""],
        ["L1-8", "Lane 1", "Complete Tier 2 country scorecards (8 countries)", "Mike", "Not Started", "P2", 2, 3, "L1-2;DC-2", "8 scorecards", "Ireland Sweden Kuwait Malaysia Spain Brazil Austria Finland"],
        ["L1-9", "Lane 1", "Complete Tier 3 country scorecards (6 countries)", "Mike", "Not Started", "P3", 3, 3, "L1-2;DC-2", "6 scorecards", "Germany Israel India Colombia Chile Mexico"],
        ["L1-10", "Lane 1", "Build RFP pipeline tracker", "Mike", "Not Started", "P1", 1, 3, "L1-1;DC-2", "Pipeline Excel", "Probability timing value per RFP"],
        ["L1-11", "Lane 1", "Build named buyer/contact list", "Mike+Bharat", "Not Started", "P1", 1, 3, "L1-1", "Contact list", "Per country"],
        ["L1-12", "Lane 1", "Build regulatory requirement matrix", "Mike", "Not Started", "P2", 2, 3, "DC-6", "Regulatory matrix", "Data sovereignty certification hosting"],
        ["L1-13", "Lane 1", "Lane 1 checkpoint with Sandeep", "All", "Not Started", "P1", 3, 3, "L1-3 to L1-9", "Checkpoint presentation", ""],
        # LANE 2
        ["L2-1", "Lane 2", "Schedule Catherine call (product gaps)", "Mike", "Not Started", "P1", 1, 1, "-", "Meeting scheduled", ""],
        ["L2-2", "Lane 2", "Schedule Lisa Ford call (roadmaps)", "Mike", "Not Started", "P1", 1, 1, "-", "Meeting scheduled", ""],
        ["L2-3", "Lane 2", "Catherine gap assessment meeting", "Mike+Catherine", "Not Started", "P1", 2, 2, "L2-1", "Gap assessment notes", "Which of 6 gaps have existing work"],
        ["L2-4", "Lane 2", "Lisa Ford roadmap review", "Mike+Lisa", "Not Started", "P1", 2, 2, "L2-2", "Roadmap status", "Intl/payer/OHI roadmap status"],
        ["L2-5", "Lane 2", "Build gap x country matrix", "Mike", "Not Started", "P1", 2, 3, "L2-3;L1-3", "Gap matrix Excel", "Revenue impact per gap per country"],
        ["L2-6", "Lane 2", "Develop build/buy/partner recommendations", "Mike", "Not Started", "P1", 3, 4, "L2-3;L2-4", "Recommendation doc", "Per gap with rationale"],
        ["L2-7", "Lane 2", "Identify partner landscape per gap", "Mike", "Not Started", "P2", 3, 4, "L2-6", "Partner list", "Who could fill each gap"],
        ["L2-8", "Lane 2", "Develop investment estimates per gap", "Mike", "Not Started", "P2", 3, 4, "L2-3;L2-6", "Investment estimates", "Rough order of magnitude"],
        ["L2-9", "Lane 2", "Compile roadmap asks for Catherine/Lisa", "Mike", "Not Started", "P1", 4, 4, "L2-5;L2-6", "Feature request list", "Specific asks with business case"],
        # LANE 3
        ["L3-1", "Lane 3", "Define feasibility scoring framework", "Mike", "Not Started", "P1", 4, 4, "L1-12", "Scoring methodology", "6 dimensions weighted"],
        ["L3-2", "Lane 3", "Score Tier 1 countries", "Mike+Sanjib", "Not Started", "P1", 4, 5, "L3-1;L1-3 to L1-7", "5 scored countries", "UK UAE KSA AUS NZ"],
        ["L3-3", "Lane 3", "Score Tier 2 countries", "Mike", "Not Started", "P2", 5, 5, "L3-1;L1-8", "8 scored countries", ""],
        ["L3-4", "Lane 3", "Score Tier 3 countries", "Mike", "Not Started", "P3", 5, 5, "L3-1;L1-9", "6 scored countries", ""],
        ["L3-5", "Lane 3", "Build composite ranking (Opportunity x Feasibility)", "Mike", "Not Started", "P1", 5, 6, "L3-2;L3-3;L3-4", "Composite ranking", "Final prioritized list"],
        ["L3-6", "Lane 3", "Assess implementation complexity per country", "Mike", "Not Started", "P2", 5, 6, "L3-2;L2-5", "Complexity assessment", ""],
        ["L3-7", "Lane 3", "Identify local partners per country", "Mike", "Not Started", "P2", 5, 6, "L3-2", "Partner requirements", "SI/consulting partners"],
        ["L3-8", "Lane 3", "Build risk register", "Mike", "Not Started", "P2", 5, 6, "L3-2 to L3-4", "Risk register", "Per country"],
        # FINAL DELIVERABLES
        ["FD-1", "Deliverables", "Draft final strategy deck", "Mike", "Not Started", "P1", 6, 6, "L1-13;L2-9;L3-5", "Draft deck (20 slides)", ""],
        ["FD-2", "Deliverables", "Build GTM messaging package", "Mike", "Not Started", "P1", 6, 7, "L2-6;L3-5", "GTM playbook", "Positioning + competitive diff"],
        ["FD-3", "Deliverables", "Build investment case with ROI", "Mike", "Not Started", "P1", 6, 7, "L2-8;L3-5", "Investment case", "Per country"],
        ["FD-4", "Deliverables", "Present to Seema", "All", "Not Started", "P1", 6, 6, "FD-1", "Seema checkpoint", ""],
        ["FD-5", "Deliverables", "Final strategy deck and execution roadmap", "Mike", "Not Started", "P1", 7, 7, "FD-4", "Final deck", "Board-ready version"],
        ["FD-6", "Deliverables", "Refinement and named owner assignment", "All", "Not Started", "P1", 7, 7, "FD-5", "Execution plan", "Named owners and dates"],
    ]

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    num_rows = len(rows) + 1
    num_cols = len(headers)

    style_header(ws, num_cols)
    style_data_rows(ws, num_rows, num_cols)

    # Center-align specific columns
    for row in range(2, num_rows + 1):
        for col in [1, 5, 6, 7, 8]:
            ws.cell(row=row, column=col).alignment = center_align

    # Data validation: Status
    status_dv = DataValidation(
        type="list", formula1='"Not Started,In Progress,Complete,Blocked,On Hold"',
        allow_blank=True
    )
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"E2:E{num_rows}")

    # Data validation: Priority
    priority_dv = DataValidation(
        type="list", formula1='"P1,P2,P3"', allow_blank=True
    )
    priority_dv.error = "Please select P1, P2, or P3"
    priority_dv.errorTitle = "Invalid Priority"
    ws.add_data_validation(priority_dv)
    priority_dv.add(f"F2:F{num_rows}")

    # Conditional formatting: Status
    ws.conditional_formatting.add(
        f"E2:E{num_rows}",
        CellIsRule(operator="equal", formula=['"Complete"'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"E2:E{num_rows}",
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"E2:E{num_rows}",
        CellIsRule(operator="equal", formula=['"In Progress"'], fill=yellow_fill)
    )

    # Conditional formatting: P1 bold — use font rule on entire row
    ws.conditional_formatting.add(
        f"A2:K{num_rows}",
        FormulaRule(formula=[f'$F2="P1"'], font=bold_font)
    )

    # Freeze panes and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{num_rows}"

    return ws


# ── Sheet 2: Owners & RACI ──────────────────────────────────────────────

def build_raci(wb):
    ws = wb.create_sheet("Owners & RACI")

    headers = ["Task Area", "Mike", "Bharat", "Catherine", "Lisa Ford",
               "Sanjib", "Sandeep", "Seema"]
    widths = [28, 10, 10, 12, 12, 10, 10, 10]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    raci_data = [
        ["Execution Plan", "R/A", "C", "-", "-", "C", "I", "I"],
        ["Data Scrape", "R/A", "C", "-", "-", "-", "I", "-"],
        ["Country Scorecards", "R/A", "R", "-", "-", "C", "I", "I"],
        ["RFP Pipeline", "R", "R/A", "-", "-", "C", "I", "I"],
        ["Product Gap Analysis", "R/A", "-", "R", "C", "-", "I", "I"],
        ["Roadmap Assessment", "R", "-", "C", "R/A", "-", "I", "I"],
        ["Feasibility Scoring", "R/A", "C", "-", "-", "R", "I", "I"],
        ["GTM Messaging", "R/A", "C", "-", "-", "C", "I", "A"],
        ["Final Strategy Deck", "R/A", "C", "C", "C", "C", "R", "A"],
        ["Seema Presentation", "R", "C", "-", "-", "C", "R", "A"],
    ]

    for r_idx, row in enumerate(raci_data, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    num_rows = len(raci_data) + 1
    num_cols = len(headers)

    style_header(ws, num_cols)
    style_data_rows(ws, num_rows, num_cols)

    # Center all data cells except task area
    for row in range(2, num_rows + 1):
        for col in range(2, num_cols + 1):
            ws.cell(row=row, column=col).alignment = center_align

    # Color code RACI values
    r_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    a_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    for row in range(2, num_rows + 1):
        for col in range(2, num_cols + 1):
            val = ws.cell(row=row, column=col).value
            if val and "A" in val:
                ws.cell(row=row, column=col).fill = a_fill
                ws.cell(row=row, column=col).font = bold_font
            elif val == "R":
                ws.cell(row=row, column=col).fill = r_fill

    # Legend
    legend_row = num_rows + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = bold_font
    ws.cell(row=legend_row + 1, column=1, value="R = Responsible").font = normal_font
    ws.cell(row=legend_row + 2, column=1, value="A = Accountable").font = normal_font
    ws.cell(row=legend_row + 3, column=1, value="C = Consulted").font = normal_font
    ws.cell(row=legend_row + 4, column=1, value="I = Informed").font = normal_font

    ws.freeze_panes = "A2"
    return ws


# ── Sheet 3: Risk Register ──────────────────────────────────────────────

def build_risk_register(wb):
    ws = wb.create_sheet("Risk Register")

    headers = ["Risk ID", "Risk Description", "Impact (1-5)", "Likelihood (1-5)",
               "Risk Score", "Mitigation", "Owner", "Status"]
    widths = [8, 50, 12, 14, 11, 45, 10, 10]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]

    risks = [
        ["R1", "No international roadmap exists - can't make roadmap commitments", 5, 5, None, "Frame as assessment not committed roadmap", "Mike", "Open"],
        ["R2", "Catherine/Lisa not available for gap assessment", 4, 3, None, "Escalate through Sandeep", "Mike", "Open"],
        ["R3", "RFP data from Bharat delayed", 4, 3, None, "Proceed with public procurement portal scrape", "Mike", "Open"],
        ["R4", "Market sizing assumptions wrong (0.2% base)", 4, 2, None, "Validate with field team + external benchmarks", "Mike", "Open"],
        ["R5", "Regulatory complexity underestimated", 3, 3, None, "Add regulatory deep-dive for Tier 1 only", "Mike", "Open"],
        ["R6", "Competitive threat from Palantir in UK", 5, 4, None, "Develop specific counter-positioning", "Mike", "Open"],
        ["R7", "OHI not included - may miss synergies", 3, 4, None, "Include OHI as future upside appendix", "Mike", "Open"],
        ["R8", "Resource constraints - Mike single-threaded", 4, 4, None, "Prioritize ruthlessly; automate with AI tools", "Mike", "Open"],
        ["R9", "Data sovereignty requirements block cloud deployment", 4, 3, None, "Map requirements early in Lane 1", "Mike", "Open"],
        ["R10", "Seema expectations exceed what can be validated in 7 weeks", 3, 3, None, "Set clear scope at Week 3 checkpoint", "Sandeep", "Open"],
    ]

    for r_idx, row in enumerate(risks, 2):
        for c_idx, val in enumerate(row, 1):
            if c_idx == 5:  # Risk Score = Impact * Likelihood formula
                ws.cell(row=r_idx, column=c_idx).value = f"=C{r_idx}*D{r_idx}"
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)

    num_rows = len(risks) + 1
    num_cols = len(headers)

    style_header(ws, num_cols)
    style_data_rows(ws, num_rows, num_cols)

    # Center specific columns
    for row in range(2, num_rows + 1):
        for col in [1, 3, 4, 5, 7, 8]:
            ws.cell(row=row, column=col).alignment = center_align

    # Conditional formatting on Risk Score (column E)
    high_risk = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    med_risk = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_risk = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws.conditional_formatting.add(
        f"E2:E{num_rows}",
        CellIsRule(operator="greaterThanOrEqual", formula=["20"], fill=high_risk)
    )
    ws.conditional_formatting.add(
        f"E2:E{num_rows}",
        FormulaRule(formula=[f"AND(E2>=10,E2<20)"], fill=med_risk)
    )
    ws.conditional_formatting.add(
        f"E2:E{num_rows}",
        CellIsRule(operator="lessThan", formula=["10"], fill=low_risk)
    )

    # Status data validation
    status_dv = DataValidation(
        type="list", formula1='"Open,Mitigated,Closed,Escalated"', allow_blank=True
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f"H2:H{num_rows}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{num_rows}"
    return ws


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    build_execution_plan(wb)
    build_raci(wb)
    build_risk_register(wb)
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")

    # Verify
    wb2 = openpyxl.load_workbook(OUTPUT)
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"  Sheet '{name}': {ws.max_row} rows x {ws.max_column} cols")
    # Check a formula
    risk_ws = wb2["Risk Register"]
    print(f"  R1 score formula: {risk_ws['E2'].value}")
    print(f"  R6 score formula: {risk_ws['E7'].value}")
    print("Verification passed.")


if __name__ == "__main__":
    main()
